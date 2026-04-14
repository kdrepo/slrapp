import csv
import io
import os

from reviews.models import Paper, Review


def import_titles_file_for_review(review_id, uploaded_file):
    review = Review.objects.get(pk=review_id)
    rows = _read_first_column(uploaded_file)
    normalized_seen = set()
    existing = {
        _normalize_title(p.title): p.id
        for p in review.papers.only('id', 'title')
        if _normalize_title(p.title)
    }

    created = 0
    duplicates = 0
    empty = 0
    samples = []

    for idx, raw_title in enumerate(rows, start=1):
        title = str(raw_title or '').strip()
        if not title:
            empty += 1
            continue

        norm = _normalize_title(title)
        if not norm:
            empty += 1
            continue

        if _looks_like_header(title, idx):
            continue

        if norm in normalized_seen or norm in existing:
            duplicates += 1
            continue

        paper = Paper.objects.create(
            review=review,
            title=title,
            abstract='',
            ta_decision=Paper.TADecision.NOT_PROCESSED,
            title_screening_decision=Paper.TitleScreeningDecision.NOT_PROCESSED,
        )
        normalized_seen.add(norm)
        existing[norm] = paper.id
        created += 1
        if len(samples) < 10:
            samples.append({'paper_id': paper.id, 'title': title})

    return {
        'file_name': uploaded_file.name,
        'total_rows': len(rows),
        'created': created,
        'duplicates_skipped': duplicates,
        'empty_rows': empty,
        'samples': samples,
    }


def _read_first_column(uploaded_file):
    ext = os.path.splitext(uploaded_file.name or '')[1].lower()

    if ext == '.csv':
        data = uploaded_file.read()
        if isinstance(data, bytes):
            text = data.decode('utf-8', errors='ignore')
        else:
            text = str(data)
        reader = csv.reader(io.StringIO(text))
        return [row[0] if row else '' for row in reader]

    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError('openpyxl is required for .xlsx upload. Install with: pip install openpyxl') from exc

        file_obj = io.BytesIO(uploaded_file.read())
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        values = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            values.append(row[0] if row else '')
        return values

    raise RuntimeError('Unsupported file type. Upload .csv or .xlsx with titles in first column.')


def _normalize_title(value):
    return ' '.join(str(value or '').strip().lower().split())


def _looks_like_header(value, row_number):
    if row_number != 1:
        return False
    text = str(value or '').strip().lower()
    return text in {'title', 'paper title', 'titles', 'paper_titles'}
