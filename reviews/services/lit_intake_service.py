import json
import os
import re
import shutil
import time
import uuid
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

from django.conf import settings

from reviews.models import LitPaper, LitReview


def ingest_ris_for_lit_review(*, review_id, file_path):
    review = LitReview.objects.get(pk=review_id)
    entries = _load_ris_entries(file_path)
    created = 0
    skipped = 0

    for entry in entries:
        title = _safe_text(entry.get('title'))[:500]
        if not title:
            skipped += 1
            continue

        doi = _safe_text(entry.get('doi'))[:255]
        url = _first_url(entry.get('urls'))
        pdf_link = url if _looks_like_pdf_url(url) else ''
        year = _parse_int(entry.get('year'))
        source = _safe_text(entry.get('secondary_title'))[:255]
        authors = _join_values(entry.get('authors', []), separator='; ')

        if _paper_exists(review=review, title=title, doi=doi):
            skipped += 1
            continue

        LitPaper.objects.create(
            review=review,
            title=title,
            authors=authors,
            year=year,
            source=source,
            doi=doi,
            url=url,
            pdf_link=pdf_link,
            origin=LitPaper.Origin.RIS_UPLOAD,
        )
        created += 1

    return {'created': created, 'skipped': skipped, 'total_rows': len(entries)}


def ingest_excel_for_lit_review(*, review_id, file_path):
    review = LitReview.objects.get(pk=review_id)
    rows = _load_excel_rows(file_path)
    created = 0
    skipped = 0

    for idx, row in enumerate(rows, start=1):
        title = _safe_text(row.get('title'))[:500]
        if not title:
            skipped += 1
            continue

        pdf_link = _safe_text(row.get('pdf_link'))
        if _paper_exists(review=review, title=title, doi=''):
            skipped += 1
            continue

        LitPaper.objects.create(
            review=review,
            title=title,
            pdf_link=pdf_link,
            origin=LitPaper.Origin.EXCEL_UPLOAD,
            excel_row_index=idx,
        )
        created += 1

    return {'created': created, 'skipped': skipped, 'total_rows': len(rows)}


def attach_numbered_pdfs_for_lit_review(*, review_id, uploaded_files):
    review = LitReview.objects.get(pk=review_id)
    matched = 0
    unmatched = 0
    errors = 0
    rows = []

    excel_papers = {
        paper.excel_row_index: paper
        for paper in review.papers.filter(origin=LitPaper.Origin.EXCEL_UPLOAD, excel_row_index__isnull=False)
    }

    for uploaded in uploaded_files:
        filename = uploaded.name or ''
        number = _extract_numbered_pdf_index(filename)
        if number is None:
            unmatched += 1
            rows.append({'file': filename, 'status': 'unmatched', 'reason': 'Filename is not in N.pdf format.'})
            continue

        paper = excel_papers.get(number)
        if not paper:
            unmatched += 1
            rows.append({'file': filename, 'status': 'unmatched', 'reason': f'No Excel row mapped for index {number}.'})
            continue

        if not filename.lower().endswith('.pdf'):
            errors += 1
            rows.append({'file': filename, 'status': 'error', 'reason': 'Not a PDF file.'})
            continue

        relative_path = _save_uploaded_pdf(review_id=review.id, paper_id=paper.id, uploaded_file=uploaded)
        paper.pdf_path.name = relative_path
        paper.fulltext_retrieved = True
        paper.pdf_source = 'numbered_upload'
        paper.save(update_fields=['pdf_path', 'fulltext_retrieved', 'pdf_source'])
        matched += 1
        rows.append({'file': filename, 'status': 'matched', 'paper_id': paper.id, 'title': paper.title})

    return {'matched': matched, 'unmatched': unmatched, 'errors': errors, 'rows': rows}


def download_missing_pdfs_for_lit_review(*, review_id):
    review = LitReview.objects.get(pk=review_id)
    papers = review.papers.filter(fulltext_retrieved=False).order_by('id')
    downloaded = 0
    skipped = 0
    failed = 0
    rows = []

    for paper in papers:
        link = (paper.pdf_link or paper.url or '').strip()
        if not link:
            skipped += 1
            rows.append({'paper_id': paper.id, 'title': paper.title, 'status': 'skipped', 'reason': 'No link available.'})
            continue

        result = _fetch_pdf_bytes(link)
        if not result.get('ok'):
            failed += 1
            rows.append({'paper_id': paper.id, 'title': paper.title, 'status': 'failed', 'reason': result.get('error', 'Unknown error')})
            continue

        relative_path = _save_pdf_bytes(review_id=review.id, paper_id=paper.id, pdf_bytes=result['pdf_bytes'])
        paper.pdf_path.name = relative_path
        paper.fulltext_retrieved = True
        paper.pdf_source = 'direct_link'
        paper.save(update_fields=['pdf_path', 'fulltext_retrieved', 'pdf_source'])
        downloaded += 1
        rows.append({'paper_id': paper.id, 'title': paper.title, 'status': 'downloaded'})

    return {'downloaded': downloaded, 'skipped': skipped, 'failed': failed, 'rows': rows, 'targeted': papers.count()}


def resolve_and_download_missing_pdfs_for_lit_review(*, review_id, progress_callback=None, stop_check=None):
    review = LitReview.objects.get(pk=review_id)
    papers = list(review.papers.filter(fulltext_retrieved=False).order_by('id'))
    downloaded = 0
    failed = 0
    resolved = 0
    rows = []

    delay_seconds = float(getattr(settings, 'PDF_RETRIEVAL_DELAY_SECONDS', 1.0))
    processed_ids = []

    _emit(
        progress_callback,
        {
            'event': 'started',
            'targeted': len(papers),
            'paper_ids': [paper.id for paper in papers],
        },
    )

    for index, paper in enumerate(papers, start=1):
        if stop_check and stop_check():
            _emit(progress_callback, {'event': 'stopped', 'processed_ids': list(processed_ids)})
            break

        note = []
        _emit(
            progress_callback,
            {
                'event': 'processing',
                'paper_id': paper.id,
                'title': paper.title,
                'index': index,
                'targeted': len(papers),
            },
        )

        # 1) Resolve DOI/Open links if missing
        if not (paper.doi or '').strip():
            scopus = _search_scopus_by_title(paper.title)
            if scopus.get('doi'):
                paper.doi = scopus['doi']
                resolved += 1
                note.append('DOI from Scopus')
            if scopus.get('url') and not paper.url:
                paper.url = scopus['url']
                note.append('URL from Scopus')

        if not (paper.doi or '').strip():
            s2 = _search_semantic_scholar_by_title(paper.title)
            if s2.get('doi'):
                paper.doi = s2['doi']
                resolved += 1
                note.append('DOI from Semantic Scholar')
            if s2.get('pdf_url') and not paper.pdf_link:
                paper.pdf_link = s2['pdf_url']
                resolved += 1
                note.append('OA PDF from Semantic Scholar')
            if s2.get('url') and not paper.url:
                paper.url = s2['url']
                note.append('URL from Semantic Scholar')

        if not (paper.doi or '').strip():
            cr = _search_crossref_by_title(paper.title)
            if cr.get('doi'):
                paper.doi = cr['doi']
                resolved += 1
                note.append('DOI from Crossref')
            if cr.get('url') and not paper.url:
                paper.url = cr['url']
                note.append('URL from Crossref')

        paper.save(update_fields=['doi', 'pdf_link', 'url'])

        # 2) Download attempt priority: direct link -> elsevier(unlocked by DOI) -> unpaywall(DOI)
        downloaded_now = False
        direct_link = (paper.pdf_link or paper.url or '').strip()
        if direct_link:
            link_result = _fetch_pdf_bytes(direct_link)
            if link_result.get('ok'):
                relative_path = _save_pdf_bytes(review_id=review.id, paper_id=paper.id, pdf_bytes=link_result['pdf_bytes'])
                paper.pdf_path.name = relative_path
                paper.fulltext_retrieved = True
                paper.pdf_source = 'resolved_direct_link'
                paper.save(update_fields=['pdf_path', 'fulltext_retrieved', 'pdf_source'])
                downloaded += 1
                downloaded_now = True
                rows.append({'paper_id': paper.id, 'title': paper.title, 'status': 'downloaded', 'source': 'direct_link', 'note': '; '.join(note)})

        if downloaded_now:
            processed_ids.append(paper.id)
            time.sleep(max(delay_seconds, 0.0))
            _emit(progress_callback, {'event': 'done', 'paper_id': paper.id, 'title': paper.title, 'source': 'direct_link'})
            continue

        doi = (paper.doi or '').strip()
        if doi:
            elsevier = _try_elsevier_pdf(doi=doi, delay_seconds=delay_seconds)
            if elsevier.get('ok'):
                relative_path = _save_pdf_bytes(review_id=review.id, paper_id=paper.id, pdf_bytes=elsevier['pdf_bytes'])
                paper.pdf_path.name = relative_path
                paper.fulltext_retrieved = True
                paper.pdf_source = 'resolved_elsevier'
                paper.save(update_fields=['pdf_path', 'fulltext_retrieved', 'pdf_source'])
                downloaded += 1
                processed_ids.append(paper.id)
                rows.append({'paper_id': paper.id, 'title': paper.title, 'status': 'downloaded', 'source': 'elsevier', 'note': '; '.join(note)})
                _emit(progress_callback, {'event': 'done', 'paper_id': paper.id, 'title': paper.title, 'source': 'elsevier'})
                continue

            unpaywall = _try_unpaywall_pdf(doi=doi, delay_seconds=delay_seconds)
            if unpaywall.get('ok'):
                relative_path = _save_pdf_bytes(review_id=review.id, paper_id=paper.id, pdf_bytes=unpaywall['pdf_bytes'])
                paper.pdf_path.name = relative_path
                paper.fulltext_retrieved = True
                paper.pdf_source = 'resolved_unpaywall'
                paper.save(update_fields=['pdf_path', 'fulltext_retrieved', 'pdf_source'])
                downloaded += 1
                processed_ids.append(paper.id)
                rows.append({'paper_id': paper.id, 'title': paper.title, 'status': 'downloaded', 'source': 'unpaywall', 'note': '; '.join(note)})
                _emit(progress_callback, {'event': 'done', 'paper_id': paper.id, 'title': paper.title, 'source': 'unpaywall'})
                continue

        failed += 1
        processed_ids.append(paper.id)
        rows.append({'paper_id': paper.id, 'title': paper.title, 'status': 'failed', 'source': 'resolver', 'note': '; '.join(note)})
        _emit(progress_callback, {'event': 'failed', 'paper_id': paper.id, 'title': paper.title, 'error_message': '; '.join(note) or 'No downloadable source found.'})

    summary = {
        'targeted': len(papers),
        'resolved': resolved,
        'downloaded': downloaded,
        'failed': failed,
        'rows': rows,
        'processed_ids': processed_ids,
        'remaining_paper_ids': [paper.id for paper in papers if paper.id not in set(processed_ids)],
        'stopped': bool(stop_check and stop_check()),
    }
    _emit(progress_callback, {'event': 'completed', **summary})
    return summary


def stage_and_extract_titles_from_uploaded_pdfs_for_lit_review(*, review_id, uploaded_files):
    review = LitReview.objects.get(pk=review_id)
    stage_dir = _title_extract_stage_dir(review.id)
    stage_dir.mkdir(parents=True, exist_ok=True)

    rows = []
    for uploaded in uploaded_files:
        file_name = uploaded.name or 'upload.pdf'
        if not file_name.lower().endswith('.pdf'):
            rows.append(
                {
                    'row_id': str(uuid.uuid4()),
                    'original_name': file_name,
                    'staged_relative_path': '',
                    'extracted_title': '',
                    'status': 'error',
                    'error': 'File is not a PDF.',
                }
            )
            continue

        staged_name = _unique_staged_name(file_name=file_name)
        staged_abs = stage_dir / staged_name
        with open(staged_abs, 'wb') as handle:
            for chunk in uploaded.chunks():
                handle.write(chunk)

        try:
            extracted_title = _extract_pdf_title_from_path(staged_abs, file_name)
            rows.append(
                {
                    'row_id': str(uuid.uuid4()),
                    'original_name': file_name,
                    'staged_relative_path': str(Path('lit_pdfs') / str(review.id) / 'title_extract_staging' / staged_name).replace('\\', '/'),
                    'extracted_title': extracted_title,
                    'status': 'ready',
                    'error': '',
                }
            )
        except Exception as exc:
            rows.append(
                {
                    'row_id': str(uuid.uuid4()),
                    'original_name': file_name,
                    'staged_relative_path': str(Path('lit_pdfs') / str(review.id) / 'title_extract_staging' / staged_name).replace('\\', '/'),
                    'extracted_title': '',
                    'status': 'error',
                    'error': f'Title extraction failed: {exc.__class__.__name__}: {exc}',
                }
            )

    return {
        'review_id': review.id,
        'rows': rows,
        'total_files': len(list(uploaded_files)),
        'ready_count': len([row for row in rows if row.get('status') == 'ready']),
        'error_count': len([row for row in rows if row.get('status') == 'error']),
    }


def finalize_verified_title_extract_rows_for_lit_review(*, review_id, rows):
    review = LitReview.objects.get(pk=review_id)
    created = 0
    skipped = 0
    errors = 0
    results = []

    for row in rows:
        include = bool(row.get('include'))
        staged_relative_path = _safe_text(row.get('staged_relative_path'))
        verified_title = _safe_text(row.get('verified_title'))[:500]
        original_name = _safe_text(row.get('original_name'))
        row_id = _safe_text(row.get('row_id'))

        if not include:
            skipped += 1
            results.append({'row_id': row_id, 'status': 'skipped', 'reason': 'User excluded row.'})
            continue

        if not verified_title:
            errors += 1
            results.append({'row_id': row_id, 'status': 'error', 'reason': 'Verified title is empty.'})
            continue

        staged_abs = _absolute_staged_path_for_review(review_id=review.id, staged_relative_path=staged_relative_path)
        if not staged_abs or not staged_abs.exists():
            errors += 1
            results.append({'row_id': row_id, 'status': 'error', 'reason': 'Staged PDF file not found.'})
            continue

        if _paper_exists(review=review, title=verified_title, doi=''):
            skipped += 1
            try:
                staged_abs.unlink(missing_ok=True)
            except Exception:
                pass
            results.append({'row_id': row_id, 'status': 'skipped', 'reason': 'Duplicate title already exists.'})
            continue

        paper = LitPaper.objects.create(
            review=review,
            title=verified_title,
            origin=LitPaper.Origin.PDF_UPLOAD,
            fulltext_retrieved=True,
            pdf_source='title_verified_upload',
        )

        final_relative = _save_pdf_from_path_for_lit_paper(review_id=review.id, paper_id=paper.id, source_path=staged_abs)
        paper.pdf_path.name = final_relative
        paper.save(update_fields=['pdf_path'])

        created += 1
        results.append({'row_id': row_id, 'status': 'created', 'paper_id': paper.id, 'title': paper.title, 'original_name': original_name})

        try:
            staged_abs.unlink(missing_ok=True)
        except Exception:
            pass

    return {'created': created, 'skipped': skipped, 'errors': errors, 'rows': results}


def _load_ris_entries(file_path):
    try:
        import rispy
    except ImportError as exc:
        raise RuntimeError('rispy is not installed. Install with: pip install rispy') from exc

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as handle:
        data = rispy.load(handle)
    return data or []


def _load_excel_rows(file_path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is not installed. Install with: pip install openpyxl') from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []

        columns = [_safe_text(cell) for cell in header]
        title_idx = _find_column(columns, {'title', 'paper_title', 'name'})
        pdf_idx = _find_column(columns, {'pdf_link', 'pdf link', 'pdf url', 'pdf_url', 'link', 'url'})
        if title_idx is None or pdf_idx is None:
            normalized_headers = [_normalize_header_name(col) for col in columns if _normalize_header_name(col)]
            raise RuntimeError(
                'Excel must contain columns for title and pdf_link. '
                f'Detected headers: {normalized_headers}'
            )

        rows = []
        for row in rows_iter:
            title = _safe_text(row[title_idx] if title_idx < len(row) else '')
            pdf_link = _safe_text(row[pdf_idx] if pdf_idx < len(row) else '')
            rows.append({'title': title, 'pdf_link': pdf_link})
        return rows
    finally:
        workbook.close()


def _find_column(columns, accepted):
    accepted_normalized = {_normalize_header_name(item) for item in accepted if _normalize_header_name(item)}
    for idx, name in enumerate(columns):
        norm = _normalize_header_name(name)
        if norm in accepted_normalized:
            return idx
    return None


def _normalize_header_name(value):
    text = _safe_text(value)
    if not text:
        return ''
    text = text.replace('\ufeff', '')
    text = re.sub(r'[\u200b-\u200d\u2060]', '', text)
    text = text.strip().lower()
    text = text.replace('_', ' ')
    text = re.sub(r'[^a-z0-9 ]+', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _paper_exists(*, review, title, doi):
    qs = review.papers.all()
    if doi:
        if qs.filter(doi__iexact=doi).exists():
            return True
    normalized = ' '.join(title.lower().split())
    for existing in qs.only('title'):
        current = ' '.join((existing.title or '').lower().split())
        if current and current == normalized:
            return True
    return False


def _extract_numbered_pdf_index(filename):
    match = re.match(r'^\s*(\d+)\.pdf\s*$', (filename or '').lower())
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _save_uploaded_pdf(*, review_id, paper_id, uploaded_file):
    target_dir = Path(settings.MEDIA_ROOT) / 'lit_pdfs' / str(review_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    absolute_path = target_dir / f'{paper_id}.pdf'

    with open(absolute_path, 'wb') as handle:
        for chunk in uploaded_file.chunks():
            handle.write(chunk)

    return str(Path('lit_pdfs') / str(review_id) / f'{paper_id}.pdf').replace('\\', '/')


def _save_pdf_bytes(*, review_id, paper_id, pdf_bytes):
    target_dir = Path(settings.MEDIA_ROOT) / 'lit_pdfs' / str(review_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    absolute_path = target_dir / f'{paper_id}.pdf'
    with open(absolute_path, 'wb') as handle:
        handle.write(pdf_bytes)
    return str(Path('lit_pdfs') / str(review_id) / f'{paper_id}.pdf').replace('\\', '/')


def _fetch_pdf_bytes(url):
    request = Request(url=url, headers={'Accept': 'application/pdf'}, method='GET')
    try:
        with urlopen(request, timeout=45) as response:
            body = response.read()
            content_type = (response.headers.get('Content-Type') or '').lower()
            if b'%PDF' not in body[:1024] and 'pdf' not in content_type:
                return {'ok': False, 'error': f'Non-PDF response Content-Type: {content_type}'}
            return {'ok': True, 'pdf_bytes': body}
    except HTTPError as exc:
        return {'ok': False, 'error': f'HTTPError: {exc.code}'}
    except URLError as exc:
        return {'ok': False, 'error': f'URLError: {exc}'}
    except TimeoutError:
        return {'ok': False, 'error': 'TimeoutError'}
    except Exception as exc:
        return {'ok': False, 'error': f'{exc.__class__.__name__}: {exc}'}


def _search_semantic_scholar_by_title(title):
    query = (title or '').strip()
    if not query:
        return {}

    headers = {'Accept': 'application/json'}
    api_key = (getattr(settings, 'S2_API_KEY', '') or os.getenv('S2_API_KEY', '')).strip()
    if api_key:
        headers['x-api-key'] = api_key

    fields = 'title,externalIds,openAccessPdf,url,year'
    encoded = quote(query)
    url = f'https://api.semanticscholar.org/graph/v1/paper/search?query={encoded}&limit=1&fields={fields}'
    result = _fetch_json(url=url, headers=headers, delay_seconds=0.0)
    if not result.get('ok'):
        return {}

    data = result.get('payload') or {}
    items = data.get('data') or []
    if not items:
        return {}
    item = items[0] or {}
    ext = item.get('externalIds') or {}
    oa = item.get('openAccessPdf') or {}
    return {
        'doi': _safe_text(ext.get('DOI')),
        'pdf_url': _safe_text(oa.get('url')),
        'url': _safe_text(item.get('url')),
    }


def _search_scopus_by_title(title):
    api_key = (getattr(settings, 'ELSEVIER_API_KEY', '') or '').strip()
    if not api_key:
        return {}

    query = (title or '').strip()
    if not query:
        return {}

    encoded = quote(query, safe='')
    url = (
        'https://api.elsevier.com/content/search/scopus'
        f'?query=TITLE%28%22{encoded}%22%29&count=1&field=dc:identifier,prism:doi,prism:url'
    )
    headers = {'Accept': 'application/json', 'X-ELS-APIKey': api_key}
    result = _fetch_json(url=url, headers=headers, delay_seconds=0.0)
    if not result.get('ok'):
        return {}

    payload = result.get('payload') or {}
    search_results = payload.get('search-results') or {}
    entries = search_results.get('entry') or []
    if not entries:
        return {}

    entry = entries[0] or {}
    doi = _safe_text(entry.get('prism:doi'))
    item_url = _safe_text(entry.get('prism:url'))
    return {'doi': doi, 'url': item_url}


def _search_crossref_by_title(title):
    query = (title or '').strip()
    if not query:
        return {}
    encoded = quote(query)
    url = f'https://api.crossref.org/works?query.title={encoded}&rows=1'
    result = _fetch_json(url=url, headers={'Accept': 'application/json'}, delay_seconds=0.0)
    if not result.get('ok'):
        return {}

    payload = result.get('payload') or {}
    message = payload.get('message') or {}
    items = message.get('items') or []
    if not items:
        return {}
    item = items[0] or {}
    doi = _safe_text(item.get('DOI'))
    item_url = _safe_text(item.get('URL'))
    return {'doi': doi, 'url': item_url}


def _try_elsevier_pdf(doi, delay_seconds):
    api_key = (getattr(settings, 'ELSEVIER_API_KEY', '') or '').strip()
    if not api_key:
        return {'ok': False, 'status_code': None, 'error_message': 'ELSEVIER_API_KEY is missing.'}

    encoded_doi = quote(doi, safe='')
    url = f'https://api.elsevier.com/content/article/doi/{encoded_doi}?httpAccept=application/pdf'
    headers = {'X-ELS-APIKey': api_key, 'Accept': 'application/pdf'}
    return _fetch_pdf_bytes_with_delay(url=url, headers=headers, delay_seconds=delay_seconds)


def _try_unpaywall_pdf(doi, delay_seconds):
    email = (getattr(settings, 'UNPAYWALL_EMAIL', '') or '').strip()
    if not email:
        return {'ok': False, 'status_code': None, 'error_message': 'UNPAYWALL_EMAIL is missing.'}

    encoded_doi = quote(doi, safe='')
    meta_url = f'https://api.unpaywall.org/v2/{encoded_doi}?email={quote(email, safe="@.")}'
    meta_result = _fetch_json(url=meta_url, headers={'Accept': 'application/json'}, delay_seconds=delay_seconds)
    if not meta_result.get('ok'):
        return {'ok': False, 'status_code': meta_result.get('status_code'), 'error_message': meta_result.get('error_message', '')}

    payload = meta_result.get('payload') or {}
    locations = []
    best_location = payload.get('best_oa_location')
    if isinstance(best_location, dict):
        locations.append(best_location)
    if isinstance(payload.get('oa_locations'), list):
        locations.extend([loc for loc in payload.get('oa_locations') if isinstance(loc, dict)])

    for location in locations:
        pdf_url = (location.get('url_for_pdf') or location.get('url') or '').strip()
        if not pdf_url:
            continue
        pdf_result = _fetch_pdf_bytes_with_delay(url=pdf_url, headers={'Accept': 'application/pdf'}, delay_seconds=delay_seconds)
        if pdf_result.get('ok'):
            return pdf_result

    return {'ok': False, 'status_code': 200, 'error_message': 'No downloadable OA PDF URL from Unpaywall.'}


def _fetch_json(url, headers, delay_seconds):
    time.sleep(max(delay_seconds, 0.0))
    request = Request(url=url, headers=headers or {}, method='GET')
    try:
        with urlopen(request, timeout=30) as response:
            raw = response.read().decode('utf-8', errors='ignore')
            return {'ok': True, 'status_code': response.getcode(), 'payload': json.loads(raw)}
    except HTTPError as exc:
        return {'ok': False, 'status_code': exc.code, 'error_message': f'HTTPError: {exc}'}
    except URLError as exc:
        return {'ok': False, 'status_code': None, 'error_message': f'URLError: {exc}'}
    except TimeoutError as exc:
        return {'ok': False, 'status_code': None, 'error_message': f'TimeoutError: {exc}'}
    except json.JSONDecodeError as exc:
        return {'ok': False, 'status_code': None, 'error_message': f'JSONDecodeError: {exc}'}
    except Exception as exc:
        return {'ok': False, 'status_code': None, 'error_message': f'UnexpectedError: {exc.__class__.__name__}: {exc}'}


def _fetch_pdf_bytes_with_delay(url, headers, delay_seconds):
    time.sleep(max(delay_seconds, 0.0))
    request = Request(url=url, headers=headers or {}, method='GET')
    try:
        with urlopen(request, timeout=45) as response:
            content_type = (response.headers.get('Content-Type') or '').lower()
            body = response.read()
            if b'%PDF' not in body[:1024] and 'pdf' not in content_type:
                return {
                    'ok': False,
                    'status_code': response.getcode(),
                    'error_message': f'Non-PDF response Content-Type: {content_type}',
                }
            return {'ok': True, 'status_code': response.getcode(), 'pdf_bytes': body}
    except HTTPError as exc:
        return {'ok': False, 'status_code': exc.code, 'error_message': f'HTTPError: {exc}'}
    except URLError as exc:
        return {'ok': False, 'status_code': None, 'error_message': f'URLError: {exc}'}
    except TimeoutError as exc:
        return {'ok': False, 'status_code': None, 'error_message': f'TimeoutError: {exc}'}
    except Exception as exc:
        return {'ok': False, 'status_code': None, 'error_message': f'UnexpectedError: {exc.__class__.__name__}: {exc}'}


def _looks_like_pdf_url(value):
    text = _safe_text(value).lower()
    return text.endswith('.pdf') or '/pdf' in text


def _first_url(urls):
    if isinstance(urls, list):
        return _safe_text(urls[0]) if urls else ''
    return _safe_text(urls)


def _parse_int(value):
    text = _safe_text(value)
    if not text:
        return None
    match = re.search(r'\d{4}', text)
    if not match:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def _join_values(value, separator='; '):
    if isinstance(value, list):
        return separator.join(_safe_text(item) for item in value if _safe_text(item))
    return _safe_text(value)


def _safe_text(value):
    if value is None:
        return ''
    return str(value).strip()


def _emit(callback, payload):
    if not callback:
        return
    try:
        callback(payload)
    except Exception:
        return


def _title_extract_stage_dir(review_id):
    return Path(settings.MEDIA_ROOT) / 'lit_pdfs' / str(review_id) / 'title_extract_staging'


def _unique_staged_name(*, file_name):
    safe = re.sub(r'[^A-Za-z0-9._-]+', '_', file_name or 'upload.pdf').strip('._')
    if not safe:
        safe = 'upload.pdf'
    stem, ext = os.path.splitext(safe)
    ext = ext or '.pdf'
    token = uuid.uuid4().hex[:10]
    return f'{stem}_{token}{ext}'


def _extract_pdf_title_from_path(path_obj, original_name):
    with open(path_obj, 'rb') as handle:
        raw = handle.read()

    if not raw:
        return ''

    text = raw.decode('latin-1', errors='ignore')

    metadata_match = re.search(r'/Title\s*\((.{5,500}?)\)', text, flags=re.IGNORECASE | re.DOTALL)
    if metadata_match:
        return _clean_extracted_title(metadata_match.group(1))

    xmp_match = re.search(r'<dc:title>.*?<rdf:li[^>]*>(.{5,500}?)</rdf:li>.*?</dc:title>', text, flags=re.IGNORECASE | re.DOTALL)
    if xmp_match:
        return _clean_extracted_title(xmp_match.group(1))

    stem = os.path.splitext(original_name or '')[0]
    stem = stem.replace('_', ' ').replace('-', ' ').strip()
    return _clean_extracted_title(stem)


def _clean_extracted_title(value):
    cleaned = (value or '').replace('\\(', '(').replace('\\)', ')')
    cleaned = cleaned.replace('\\n', ' ').replace('\n', ' ')
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned[:500]


def _absolute_staged_path_for_review(*, review_id, staged_relative_path):
    relative = Path((staged_relative_path or '').replace('\\', '/'))
    if not relative.parts:
        return None
    stage_root = _title_extract_stage_dir(review_id).resolve()
    candidate = (Path(settings.MEDIA_ROOT) / relative).resolve()
    try:
        candidate.relative_to(stage_root)
    except Exception:
        return None
    return candidate


def _save_pdf_from_path_for_lit_paper(*, review_id, paper_id, source_path):
    target_dir = Path(settings.MEDIA_ROOT) / 'lit_pdfs' / str(review_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    final_abs = target_dir / f'{paper_id}.pdf'
    shutil.copyfile(source_path, final_abs)
    return str(Path('lit_pdfs') / str(review_id) / f'{paper_id}.pdf').replace('\\', '/')
